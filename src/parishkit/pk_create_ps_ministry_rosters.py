"""Implementation for the pk-create-ps-ministry-rosters command."""

from __future__ import annotations

import argparse
import datetime as dt
import logging
import tempfile
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from importlib.metadata import version
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from parishkit.cli import (
    parser_with_common_options,
    require_explicit_write_mode,
    resolve_common_options,
    run_user_facing,
)
from parishkit.config import (
    ConfigData,
    ConfigError,
    load_yaml_config,
    reject_unknown_keys,
    resolve_path,
)
from parishkit.google.auth import (
    GoogleAPIError,
    load_service_account_credentials,
    load_user_credentials,
)
from parishkit.google.drive import (
    GOOGLE_SHEET_MIME_TYPE,
    build_drive_service,
    get_file_metadata,
    update_file_with_xlsx,
)
from parishkit.logging import log_extra, setup_logging
from parishkit.parishsoft import (
    ParishSoftData,
    get_member_public_email,
    get_member_public_phones,
    load_families_and_members,
)
from parishkit.parishsoft_runtime import parishsoft_client_from_config

DRIVE_SCOPE = "https://www.googleapis.com/auth/drive"
DEFAULT_LEADER_SUFFIX = " Ldr"
HEADER_BACKGROUND_COLOR = "0000FF"
HEADER_TEXT_COLOR = "FFFF00"
ROSTER_COLUMN_WIDTHS = (30, 30, 50, 30, 20)
ROSTER_WORKSHEET_TITLE = "Sheet"
ROSTER_TITLE_ROWS = 2
ROSTER_TITLE_MERGE_COLUMNS = 4
ROSTER_FROZEN_ROWS = 4
ROSTER_SPACER_ROW_INDEX = 2
ROSTER_COLUMN_HEADER_ROW_INDEX = 3
DRIVE_PREFLIGHT_FIELDS = (
    "id,name,mimeType,capabilities/canEdit,capabilities/canModifyContent,"
    "capabilities/canRename"
)


@dataclass(frozen=True)
class RoleRosterTarget:
    """A secondary sheet holding only members whose role is in ``roles``.

    Each role sheet is a filtered view of its parent ministry roster, written
    to its own Drive file.
    """

    name: str
    roles: tuple[str, ...]
    spreadsheet_id: str


@dataclass(frozen=True)
class RosterTarget:
    """A configured roster to publish to a Google Drive spreadsheet file.

    ``source_type`` is either ``"ministry"`` or ``"workgroup"`` and
    ``source_names`` lists the ParishSoft sources whose members populate the
    roster. ``role_sheets`` are optional per-role breakout sheets derived from
    the same members.
    """

    name: str
    source_type: str
    source_names: tuple[str, ...]
    spreadsheet_id: str
    include_birthday: bool
    role_sheets: tuple[RoleRosterTarget, ...] = ()


@dataclass(frozen=True)
class RosterConfig:
    """Parsed ``rosters`` configuration.

    ``workgroup_leader_suffix`` is appended to a workgroup name to recognize
    the companion leader workgroup in ParishSoft.
    """

    ministries: tuple[RosterTarget, ...]
    workgroups: tuple[RosterTarget, ...]
    workgroup_leader_suffix: str = DEFAULT_LEADER_SUFFIX


@dataclass(frozen=True)
class RosterMember:
    """A member paired with the role text to show on a roster row."""

    member: dict[str, Any]
    role: str


@dataclass(frozen=True)
class RosterWritePlan:
    """A validated Drive upload that is safe to apply after all plans exist."""

    spreadsheet_id: str
    title: str
    values: list[list[Any]]
    update_time: dt.datetime


@dataclass(frozen=True)
class PreparedRosterUpload:
    """A serialized workbook ready for Drive upload."""

    plan: RosterWritePlan
    drive_name: str
    xlsx_path: Path


Loader = Callable[..., ParishSoftData]
DriveFactory = Callable[[ConfigData], Any]


def _text_list(values: Sequence[str]) -> str:
    """Render a short list of strings for human-readable log messages."""
    return ", ".join(values) if values else "none"


def main(
    argv: Sequence[str] | None = None,
    *,
    loader: Loader = load_families_and_members,
    drive_factory: DriveFactory | None = None,
) -> int:
    """Run the command-line entry point."""
    parser = parser_with_common_options(
        "pk-create-ps-ministry-rosters",
        description="Publish ParishSoft ministry rosters to Google Drive Sheets.",
    )
    parser.add_argument(
        "--version",
        action="store_true",
        help="show that the console entry point is installed",
    )
    args = parser.parse_args(argv)
    if args.version:
        print(f"pk-create-ps-ministry-rosters {version('parishkit')}")
        return 0
    return run_user_facing(lambda: _run(args, loader, drive_factory))


def _run(
    args: argparse.Namespace,
    loader: Loader,
    drive_factory: DriveFactory | None,
) -> int:
    """Load config and data, then publish every configured roster.

    Resolves common CLI options, loads the YAML roster config, sets up
    logging (including optional Slack), builds the ParishSoft client and the
    Google Drive service, fetches active members, and uploads the rosters.
    The Drive service can be injected via ``drive_factory`` for testing;
    otherwise it is built from credentials in the config. Returns 0 on
    success. The steps are kept explicit so operational behavior remains easy
    to audit and test.
    """
    common = resolve_common_options(args)
    log = setup_logging(
        verbose=common.verbose or common.dry_run,
        debug=common.debug,
        log_file=common.log_file,
        log_dir=common.log_dir,
        logger_name="parishkit.pk_create_ps_ministry_rosters",
        slack_token_file=common.slack_token_file,
        slack_channel=common.slack_channel,
        slack_level=common.slack_log_level,
    )
    try:
        require_explicit_write_mode(common, "pk-create-ps-ministry-rosters")
        config = load_yaml_config(common.config)
        roster_config = roster_config_from_yaml(config)
        log.info(
            "Configured %s ministry roster(s) and %s workgroup roster(s)",
            len(roster_config.ministries),
            len(roster_config.workgroups),
        )
        log.debug(
            "Ministry roster targets: %s",
            _text_list([target.name for target in roster_config.ministries]),
            extra=log_extra(roster_config.ministries),
        )
        log.debug(
            "Workgroup roster targets: %s",
            _text_list([target.name for target in roster_config.workgroups]),
            extra=log_extra(roster_config.workgroups),
        )
        client = parishsoft_client_from_config(common, config)
        log.info("Loading active ParishSoft families and members")
        data = loader(client, active_only=True, parishioners_only=False)
        log.info(
            "Loaded %s member(s), %s family/families, %s ministry membership(s), "
            "and %s workgroup membership(s)",
            len(data.members),
            len(data.families),
            len(data.ministry_type_memberships),
            len(data.member_workgroup_memberships),
        )
        validate_configured_parishsoft_sources(data, roster_config)
        log.debug("Dry-run mode is %s", "enabled" if common.dry_run else "disabled")
        drive_service = None
        if not common.dry_run:
            drive_service = (
                drive_factory(config)
                if drive_factory is not None
                else build_drive_service(
                    load_drive_credentials(
                        config,
                        base_dir=common.config.parent if common.config else None,
                    )
                )
            )
        write_configured_rosters(
            drive_service,
            data,
            roster_config,
            timezone_name=common.timezone,
            dry_run=common.dry_run,
            log=log,
        )
    except ConfigError as exc:
        log.error("Configuration validation failed: %s", exc)
        raise
    log.info("Ministry roster operation completed successfully")
    return 0


def roster_config_from_yaml(config: ConfigData) -> RosterConfig:
    """Parse and validate the ``rosters`` config section.

    Reads optional top-level defaults and applies them to each ministry and
    workgroup target. Raises ``ConfigError`` on malformed values or when neither
    any ministries nor any workgroups are configured.
    """
    section = _mapping(
        config.get("rosters", {}),
        "rosters",
    )
    reject_unknown_keys(
        section,
        {
            "spreadsheet_id",
            "workgroup_leader_suffix",
            "ministries",
            "workgroups",
        },
        "rosters",
    )
    default_spreadsheet_id = section.get("spreadsheet_id")
    if default_spreadsheet_id is not None and not isinstance(
        default_spreadsheet_id, str
    ):
        raise ConfigError("rosters.spreadsheet_id must be a string")
    leader_suffix = section.get("workgroup_leader_suffix", DEFAULT_LEADER_SUFFIX)
    if not isinstance(leader_suffix, str):
        raise ConfigError("rosters.workgroup_leader_suffix must be a string")
    ministries = tuple(
        _target(
            item,
            f"rosters.ministries[{index}]",
            source_key="ministry",
            plural_source_key="ministries",
            source_type="ministry",
            default_spreadsheet_id=default_spreadsheet_id,
        )
        for index, item in enumerate(
            _list(section.get("ministries", []), "rosters.ministries")
        )
    )
    workgroups = tuple(
        _target(
            item,
            f"rosters.workgroups[{index}]",
            source_key="workgroup",
            plural_source_key=None,
            source_type="workgroup",
            default_spreadsheet_id=default_spreadsheet_id,
        )
        for index, item in enumerate(
            _list(section.get("workgroups", []), "rosters.workgroups")
        )
    )
    if not ministries and not workgroups:
        raise ConfigError("rosters must configure ministries or workgroups")
    roster_config = RosterConfig(
        ministries=ministries,
        workgroups=workgroups,
        workgroup_leader_suffix=leader_suffix,
    )
    validate_unique_roster_targets(roster_config)
    return roster_config


def load_drive_credentials(
    config: ConfigData,
    *,
    base_dir: Path | None = None,
) -> Any:
    """Load credentials for Google Drive file replacement access."""
    google = _mapping(config.get("google", {}), "google")
    reject_unknown_keys(
        google,
        {"service_account_file", "user_token_file", "delegated_subject"},
        "google",
    )
    service_account_file = google.get("service_account_file")
    user_token_file = google.get("user_token_file")
    delegated_subject = google.get("delegated_subject")
    if service_account_file and user_token_file:
        raise ConfigError(
            "google configuration must not set both service_account_file "
            "and user_token_file"
        )
    if delegated_subject is not None and not isinstance(delegated_subject, str):
        raise ConfigError("google.delegated_subject must be a string")
    if isinstance(service_account_file, str):
        return load_service_account_credentials(
            resolve_path(
                service_account_file,
                "google.service_account_file",
                base_dir=base_dir,
            ),
            scopes=[DRIVE_SCOPE],
            subject=delegated_subject,
        )
    if isinstance(user_token_file, str):
        return load_user_credentials(
            resolve_path(
                user_token_file,
                "google.user_token_file",
                base_dir=base_dir,
            ),
            scopes=[DRIVE_SCOPE],
        )
    raise ConfigError(
        "google.service_account_file or google.user_token_file is required"
    )


def write_configured_rosters(
    drive_service: Any,
    data: ParishSoftData,
    config: RosterConfig,
    *,
    timezone_name: str,
    dry_run: bool,
    log: logging.Logger,
) -> None:
    """Write every ministry, role, and workgroup roster through Drive uploads.

    For each ministry target, collects its members and writes the main roster
    plus any per-role breakout sheets (filtered to that sheet's allowed roles).
    For each workgroup target, collects members tagging leaders separately and
    writes the roster. ``dry_run`` logs intended uploads without touching
    Drive. All plans are built before the first upload so data/config failures
    still abort before any external write.
    """
    update_time = current_roster_time(timezone_name)
    plans: list[RosterWritePlan] = []
    for target in config.ministries:
        log.debug(
            "Preparing ministry roster %s from %s",
            target.name,
            _text_list(target.source_names),
            extra=log_extra(target),
        )
        members = ministry_roster_members(data, target.source_names)
        plans.append(
            roster_target_plan(
                target,
                members,
                update_time=update_time,
            )
        )
        for role_target in target.role_sheets:
            allowed_roles = set(role_target.roles)
            log.debug(
                "Preparing role roster %s from %s role(s): %s",
                role_target.name,
                target.name,
                _text_list(role_target.roles),
                extra=log_extra(role_target),
            )
            role_members = [
                member
                for member in members
                if roster_role_matches(member.role, allowed_roles)
            ]
            plans.append(
                write_plan(
                    role_target.spreadsheet_id,
                    role_target.name,
                    roster_values(
                        role_target.name,
                        role_members,
                        include_birthday=target.include_birthday,
                        now=update_time,
                    ),
                    update_time=update_time,
                )
            )
    for target in config.workgroups:
        log.debug(
            "Preparing workgroup roster %s from %s",
            target.name,
            _text_list(target.source_names),
            extra=log_extra(target),
        )
        members = workgroup_roster_members(
            data,
            target.source_names[0],
            leader_suffix=config.workgroup_leader_suffix,
        )
        plans.append(
            roster_target_plan(
                target,
                members,
                update_time=update_time,
            )
        )
    if dry_run:
        for plan in plans:
            log_dry_run_roster_plan(plan, log)
        return
    preflight_drive_roster_targets(drive_service, plans)
    with tempfile.TemporaryDirectory(prefix="parishkit-roster-") as temp_dir:
        uploads = prepare_roster_uploads(plans, Path(temp_dir))
        for upload in uploads:
            upload_prepared_roster(drive_service, upload, log=log)


def preflight_drive_roster_targets(
    drive_service: Any,
    plans: Sequence[RosterWritePlan],
) -> None:
    """Verify all Drive targets are usable before the first upload.

    Replacing a Drive file is a live write. Checking every configured target
    first avoids partially updated roster sets when a later file ID is wrong,
    inaccessible, not editable by the delegated user, or not a native Google
    spreadsheet.
    """
    for plan in plans:
        try:
            metadata = get_file_metadata(
                drive_service,
                plan.spreadsheet_id,
                fields=DRIVE_PREFLIGHT_FIELDS,
            )
        except GoogleAPIError as exc:
            raise ConfigError(
                "Could not access configured Drive roster target for "
                f"{plan.title!r}: {plan.spreadsheet_id}. Check the "
                "rosters.*.spreadsheet_id value and share the file or shared "
                "drive with the delegated Google user as Editor."
            ) from exc
        label = drive_target_label(plan, metadata)
        if metadata.get("mimeType") != GOOGLE_SHEET_MIME_TYPE:
            raise ConfigError(
                f"Configured Drive roster target for {plan.title!r} must be a "
                f"native Google Sheet: {label} has MIME type "
                f"{metadata.get('mimeType')!r}."
            )
        capabilities = metadata.get("capabilities") or {}
        if not capabilities.get("canEdit"):
            raise ConfigError(
                f"Configured Drive roster target for {plan.title!r} is not "
                f"editable by the delegated Google user: {label}. Share the "
                "file or shared drive with that user as Editor."
            )
        if not capabilities.get("canModifyContent"):
            raise ConfigError(
                f"Configured Drive roster target for {plan.title!r} cannot "
                f"have its content replaced by the delegated Google user: "
                f"{label}. Share the file or shared drive with that user as "
                "Editor and make sure Drive restrictions allow content changes."
            )
        if not capabilities.get("canRename"):
            raise ConfigError(
                f"Configured Drive roster target for {plan.title!r} cannot "
                f"be renamed by the delegated Google user: {label}. Share the "
                "file or shared drive with that user as Editor and make sure "
                "Drive restrictions allow renaming."
            )


def drive_target_label(
    plan: RosterWritePlan,
    metadata: Mapping[str, Any],
) -> str:
    """Return a human-readable Drive target label for error messages."""
    name = metadata.get("name")
    if name:
        return f"{name!r} ({plan.spreadsheet_id})"
    return plan.spreadsheet_id


def validate_unique_roster_targets(config: RosterConfig) -> None:
    """Reject multiple roster outputs targeting the same Drive file."""
    seen: dict[str, str] = {}
    named_targets = []
    for target in config.ministries:
        named_targets.append((target.name, target.spreadsheet_id))
        named_targets.extend(
            (role.name, role.spreadsheet_id) for role in target.role_sheets
        )
    named_targets.extend(
        (target.name, target.spreadsheet_id) for target in config.workgroups
    )
    for name, spreadsheet_id in named_targets:
        if spreadsheet_id in seen:
            raise ConfigError(
                "rosters outputs must not share the same Drive file: "
                f"{seen[spreadsheet_id]!r} and {name!r} both target "
                f"{spreadsheet_id}"
            )
        seen[spreadsheet_id] = name


def validate_configured_parishsoft_sources(
    data: ParishSoftData,
    config: RosterConfig,
) -> None:
    """Verify configured roster ministry/workgroup names exist in ParishSoft."""
    ministry_names = available_ministry_names(data)
    workgroup_names = available_member_workgroup_source_names(
        data,
        leader_suffix=config.workgroup_leader_suffix,
    )
    for target in config.ministries:
        for ministry in target.source_names:
            if ministry not in ministry_names:
                raise ConfigError(
                    f"Configured ParishSoft ministry was not found for roster "
                    f"{target.name!r}: {ministry!r}. Check "
                    "rosters.ministries[].ministry or "
                    "rosters.ministries[].ministries in the YAML and make sure "
                    "each name exactly matches a ParishSoft ministry. Available "
                    f"ministries: {_text_list(sorted(ministry_names))}."
                )
    for target in config.workgroups:
        for workgroup in target.source_names:
            if workgroup not in workgroup_names:
                raise ConfigError(
                    f"Configured ParishSoft member workgroup was not found for "
                    f"roster {target.name!r}: {workgroup!r}. Check "
                    "rosters.workgroups[].workgroup in the YAML and make sure "
                    "it exactly matches a ParishSoft member workgroup. "
                    "Available member workgroups: "
                    f"{_text_list(sorted(workgroup_names))}."
                )


def available_ministry_names(data: ParishSoftData) -> set[str]:
    """Return ministry names present in loaded ParishSoft data."""
    names = {
        str(item["name"])
        for item in data.ministry_type_memberships.values()
        if item.get("name")
    }
    for member in data.members.values():
        names.update(str(name) for name in member.get("py ministries", {}))
    return names


def available_member_workgroup_source_names(
    data: ParishSoftData,
    *,
    leader_suffix: str,
) -> set[str]:
    """Return member workgroup names usable as configured source names."""
    names = {
        str(item["name"])
        for item in data.member_workgroup_memberships.values()
        if item.get("name")
    }
    for member in data.members.values():
        names.update(str(name) for name in member.get("py workgroups", {}))
    for name in tuple(names):
        for suffix in (leader_suffix, " Ldr", " Leader"):
            if suffix and name.endswith(suffix):
                names.add(name[: -len(suffix)])
    return names


def roster_target_plan(
    target: RosterTarget,
    members: Sequence[RosterMember],
    *,
    update_time: dt.datetime,
) -> RosterWritePlan:
    """Build a validated upload plan for one configured roster target."""
    return write_plan(
        target.spreadsheet_id,
        target.name,
        roster_values(
            target.name,
            members,
            include_birthday=target.include_birthday,
            now=update_time,
        ),
        update_time=update_time,
    )


def write_plan(
    spreadsheet_id: str,
    title: str,
    values: list[list[Any]],
    *,
    update_time: dt.datetime,
) -> RosterWritePlan:
    """Build one whole-workbook Drive upload plan."""
    return RosterWritePlan(
        spreadsheet_id=spreadsheet_id,
        title=title,
        values=values,
        update_time=update_time,
    )


def log_dry_run_roster_plan(plan: RosterWritePlan, log: logging.Logger) -> None:
    """Log the upload that would happen for one roster plan."""
    log.info(
        "dry-run: would upload %s row(s) for %s to Drive file %s",
        len(plan.values),
        plan.title,
        plan.spreadsheet_id,
    )


def prepare_roster_uploads(
    plans: Sequence[RosterWritePlan],
    temp_dir: Path,
) -> list[PreparedRosterUpload]:
    """Serialize every planned roster workbook before the first Drive write.

    Local workbook generation can fail on unexpected source data. Preparing the
    complete set first keeps those failures on the no-external-write side of
    the workflow, preserving the partial-update guard that Drive preflight
    provides for remote targets.
    """
    uploads = []
    for index, plan in enumerate(plans, start=1):
        drive_name = roster_drive_name(plan.title, plan.update_time)
        xlsx_path = temp_dir / f"{index:03d}-{safe_roster_filename(drive_name)}.xlsx"
        workbook = None
        try:
            workbook = roster_workbook(plan.values)
            workbook.save(xlsx_path)
        except Exception as exc:
            raise ConfigError(
                f"Could not build XLSX roster workbook for {plan.title!r}: {exc}"
            ) from exc
        finally:
            if workbook is not None:
                workbook.close()
        uploads.append(
            PreparedRosterUpload(
                plan=plan,
                drive_name=drive_name,
                xlsx_path=xlsx_path,
            )
        )
    return uploads


def upload_prepared_roster(
    drive_service: Any,
    upload: PreparedRosterUpload,
    *,
    log: logging.Logger,
) -> None:
    """Upload one already-serialized roster workbook to its Drive file."""
    update_file_with_xlsx(
        drive_service,
        upload.plan.spreadsheet_id,
        upload.xlsx_path,
        name=upload.drive_name,
    )
    log.info(
        "Uploaded roster workbook for %s to Drive file %s",
        upload.plan.title,
        upload.plan.spreadsheet_id,
    )


def roster_drive_name(title: str, update_time: dt.datetime) -> str:
    """Return the Google Drive file name for an uploaded roster workbook."""
    return f"{safe_roster_filename(title)} as of {format_update_timestamp(update_time)}"


def safe_roster_filename(value: str) -> str:
    """Return a Drive/local filename-safe roster name."""
    return value.replace("/", "-")


def roster_workbook(values: Sequence[Sequence[Any]]) -> Workbook:
    """Build an XLSX workbook containing roster values and formatting."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = ROSTER_WORKSHEET_TITLE
    column_count = max((len(row) for row in values), default=1)

    for row_index, row in enumerate(values, start=1):
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            set_roster_cell_value(cell, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    apply_roster_workbook_formatting(worksheet, column_count=column_count)
    return workbook


def set_roster_cell_value(cell: Any, value: Any) -> None:
    """Write a roster cell while preserving ParishSoft text literally.

    openpyxl treats strings beginning with ``=`` as formulas by default. Roster
    content comes from ParishSoft and should be displayed, not executed, so
    string cells are explicitly marked as string data.
    """
    cell.value = value
    if isinstance(value, str):
        cell.data_type = "s"


def apply_roster_workbook_formatting(worksheet: Any, *, column_count: int) -> None:
    """Apply the roster layout to an openpyxl worksheet."""
    header_fill = PatternFill(fgColor=HEADER_BACKGROUND_COLOR, fill_type="solid")
    header_font = Font(color=HEADER_TEXT_COLOR, bold=True)
    title_end_column = min(ROSTER_TITLE_MERGE_COLUMNS, column_count)
    title_end_letter = get_column_letter(title_end_column)
    for row_index in range(1, ROSTER_TITLE_ROWS + 1):
        worksheet.merge_cells(f"A{row_index}:{title_end_letter}{row_index}")
        cell = worksheet.cell(row=row_index, column=1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for column_index in range(1, column_count + 1):
        spacer = worksheet.cell(row=ROSTER_SPACER_ROW_INDEX + 1, column=column_index)
        spacer.fill = header_fill
        header = worksheet.cell(
            row=ROSTER_COLUMN_HEADER_ROW_INDEX + 1,
            column=column_index,
        )
        header.fill = header_fill
        header.font = header_font
        header.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in worksheet.iter_rows(min_row=ROSTER_FROZEN_ROWS + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for index, width in enumerate(ROSTER_COLUMN_WIDTHS[:column_count], start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.freeze_panes = f"A{ROSTER_FROZEN_ROWS + 1}"


def ministry_roster_members(
    data: ParishSoftData,
    ministry_names: Sequence[str],
) -> list[RosterMember]:
    """Return roster members belonging to any of ``ministry_names``.

    A member is included when they belong to at least one configured ministry.
    Their roles across the matching ministries are deduplicated, sorted, and
    joined into a single comma-separated role string. Empty roles do not appear
    in the joined text, but members with only blank roles are still included
    with an empty role cell because ParishSoft permits membership without a
    role. Results are sorted by the member sort key for stable output.
    """
    configured = set(ministry_names)
    members = []
    for member in data.members.values():
        ministries = member.get("py ministries", {})
        matching_entries = [
            entry for name, entry in ministries.items() if name in configured
        ]
        if matching_entries:
            roles = sorted(
                {str(entry.get("role") or "").strip() for entry in matching_entries}
            )
            members.append(
                RosterMember(
                    member=member,
                    role=", ".join(role for role in roles if role),
                )
            )
    return sorted(members, key=lambda item: member_sort_key(item.member))


def workgroup_roster_members(
    data: ParishSoftData,
    workgroup_name: str,
    *,
    leader_suffix: str,
) -> list[RosterMember]:
    """Return roster members for a single workgroup.

    ParishSoft models leaders as a separate companion workgroup named
    ``workgroup_name + leader_suffix``. Members in that companion group are
    labeled ``"Leader"``; members in the base workgroup are labeled
    ``"Member"``. Results are sorted by the member sort key.
    """
    leader_name = f"{workgroup_name}{leader_suffix}"
    members = []
    for member in data.members.values():
        workgroups = member.get("py workgroups", {})
        if leader_name in workgroups:
            members.append(RosterMember(member=member, role="Leader"))
        elif workgroup_name in workgroups:
            members.append(RosterMember(member=member, role="Member"))
    return sorted(members, key=lambda item: member_sort_key(item.member))


def roster_role_matches(role_text: str, allowed_roles: set[str]) -> bool:
    """Report whether any role in a comma-separated string is allowed.

    ``role_text`` is the joined role string produced for a roster member, so it
    is split on commas and each piece is stripped before comparison against
    ``allowed_roles``.
    """
    return any(role.strip() in allowed_roles for role in role_text.split(","))


def roster_values(
    title: str,
    members: Sequence[RosterMember],
    *,
    include_birthday: bool,
    now: dt.datetime | None = None,
) -> list[list[Any]]:
    """Build the 2-D cell grid for one roster sheet.

    The grid starts with a title row, a "Last updated" timestamp, a blank
    spacer row, and a header row, followed by one or more rows per member. The
    birthday column is included only when ``include_birthday`` is set. A member
    with both phone number(s) and email gets a second row so email stays
    visually separate from phone contact. ``now`` is injectable so the timestamp
    is deterministic in tests; it defaults to the current local time truncated
    to whole seconds. The steps are kept explicit so operational behavior
    remains easy to audit and test.
    """
    timestamp = format_update_timestamp(now or dt.datetime.now())
    headers = ["Member name", "Address", "Phone / email"]
    if include_birthday:
        headers.append("Birthday")
    headers.append("Role")
    rows: list[list[Any]] = [
        [f"Ministry: {title}"],
        [f"Last updated: {timestamp}"],
        [],
        headers,
    ]
    for roster_member in sorted(members, key=lambda item: member_sort_key(item.member)):
        rows.extend(
            roster_member_rows(roster_member, include_birthday=include_birthday)
        )
    return rows


def current_roster_time(timezone_name: str) -> dt.datetime:
    """Return the current update time in the configured roster timezone."""
    return dt.datetime.now(ZoneInfo(timezone_name)).replace(microsecond=0)


def format_update_timestamp(update_time: dt.datetime) -> str:
    """Format an update timestamp, including timezone abbreviation when known."""
    value = update_time.replace(microsecond=0)
    timestamp = value.strftime("%Y-%m-%d %H:%M:%S")
    timezone = value.strftime("%Z")
    if timezone:
        return f"{timestamp} {timezone}"
    return timestamp


def roster_member_rows(
    roster_member: RosterMember,
    *,
    include_birthday: bool,
) -> list[list[Any]]:
    """Build one or two output rows for a roster member.

    Phone numbers stay on the row with the member name. If an email address is
    also present, it moves to a continuation row so phone and email are not
    mixed in the same cell. When no phone number is present, email remains on
    the main row.
    """
    member = roster_member.member
    phone_text = member_phone_contact(member)
    email = get_member_public_email(member) or ""
    first_contact = phone_text or email
    first_row = [
        member.get("py friendly name LF") or member_display_name(member),
        member_address(member),
        first_contact,
    ]
    if include_birthday:
        first_row.append(member_birthday(member))
    first_row.append(roster_member.role)
    rows = [first_row]
    if phone_text and email:
        continuation_row = ["", "", email]
        if include_birthday:
            continuation_row.append("")
        continuation_row.append("")
        rows.append(continuation_row)
    return rows


def member_sort_key(member: Mapping[str, Any]) -> str:
    """Return a stable sort key for a roster member."""
    name = member.get("py friendly name LF") or member_display_name(member)
    return f"{name} {member.get('memberDUID', '')}"


def member_display_name(member: Mapping[str, Any]) -> str:
    """Return the roster display name for a member."""
    first = str(member.get("firstName") or "")
    last = str(member.get("lastName") or "")
    # Strip stray commas/spaces so a missing first or last name does not leave a
    # dangling separator (e.g. "Smith, " or ", John").
    return f"{last}, {first}".strip(", ")


def member_address(member: Mapping[str, Any]) -> str:
    """Return the roster mailing address for a member."""
    family = member.get("py family") or {}
    parts = [
        family.get("primaryAddress1"),
        family.get("primaryAddress2"),
        city_state_zip(family),
    ]
    return "\n".join(str(part) for part in parts if part)


def city_state_zip(family: Mapping[str, Any]) -> str:
    """Format a family's city, state, and postal code into one line.

    Missing parts are skipped so the result stays clean: city and state are
    comma-joined, then the postal code (if any) is appended after a space.
    """
    city = family.get("primaryCity")
    state = family.get("primaryState")
    postal_code = family.get("primaryPostalCode")
    city_state = ", ".join(str(part) for part in (city, state) if part)
    if postal_code:
        return f"{city_state} {postal_code}".strip()
    return city_state


def member_phone_contact(member: dict[str, Any]) -> str:
    """Return the roster phone text for a member."""
    return "\n".join(
        f"{phone['number']} {phone['type']}"
        for phone in get_member_public_phones(member)
    )


def member_birthday(member: Mapping[str, Any]) -> str:
    """Return the roster birthday text for a member."""
    value = member.get("birthdate")
    if isinstance(value, dt.datetime):
        value = value.date()
    if isinstance(value, dt.date):
        return f"{value.strftime('%B')} {value.day}"
    return ""


def _target(
    value: Any,
    name: str,
    *,
    source_key: str,
    plural_source_key: str | None,
    source_type: str,
    default_spreadsheet_id: str | None,
) -> RosterTarget:
    """Parse one ministry or workgroup target into a ``RosterTarget``.

    ``source_key``/``plural_source_key`` select the config keys that name the
    ParishSoft source(s); ``source_type`` records which kind this is. Per-target
    spreadsheet ID falls back to the supplied default when omitted, and the
    display name defaults to the joined source names. Both
    ``include_birthday``/``birthday`` spellings are accepted. Ministry targets
    also accept ``role_sheets``/``role sheets``.
    """
    item = _mapping(value, name)
    allowed_keys = {
        "name",
        source_key,
        "spreadsheet_id",
        "include_birthday",
        "birthday",
    }
    if plural_source_key:
        allowed_keys.add(plural_source_key)
    if source_type == "ministry":
        allowed_keys.update({"role_sheets", "role sheets"})
    reject_unknown_keys(item, allowed_keys, name)
    source_names = _source_names(item, name, source_key, plural_source_key)
    target_name = _optional_string(item.get("name"), f"{name}.name") or ", ".join(
        source_names
    )
    spreadsheet_id = _target_spreadsheet_id(item, name, default_spreadsheet_id)
    include_birthday = _bool(
        item.get("include_birthday", item.get("birthday", False)),
        f"{name}.include_birthday",
    )
    role_sheets_value = (
        item.get("role_sheets", item.get("role sheets", []))
        if source_type == "ministry"
        else []
    )
    role_sheets = tuple(
        _role_target(
            role_sheet,
            f"{name}.role_sheets[{index}]",
        )
        for index, role_sheet in enumerate(
            _list(role_sheets_value, f"{name}.role_sheets")
        )
    )
    return RosterTarget(
        name=target_name,
        source_type=source_type,
        source_names=tuple(source_names),
        spreadsheet_id=spreadsheet_id,
        include_birthday=include_birthday,
        role_sheets=role_sheets,
    )


def _role_target(value: Any, name: str) -> RoleRosterTarget:
    """Parse one role-specific roster target."""
    item = _mapping(value, name)
    reject_unknown_keys(
        item,
        {"name", "roles", "spreadsheet_id"},
        name,
    )
    role_name = _required_string(item.get("name"), f"{name}.name")
    roles = tuple(_string_list(item.get("roles"), f"{name}.roles"))
    spreadsheet_id = _required_string(
        item.get("spreadsheet_id"),
        f"{name}.spreadsheet_id",
    )
    return RoleRosterTarget(
        name=role_name,
        roles=roles,
        spreadsheet_id=spreadsheet_id,
    )


def _source_names(
    item: Mapping[str, Any],
    name: str,
    source_key: str,
    plural_source_key: str | None,
) -> list[str]:
    """Parse ministry or workgroup names from config."""
    singular = item.get(source_key)
    plural = item.get(plural_source_key) if plural_source_key else None
    if singular and plural:
        raise ConfigError(
            f"{name} must not set both {source_key} and {plural_source_key}"
        )
    if singular is not None:
        return [_required_string(singular, f"{name}.{source_key}")]
    if plural_source_key:
        return _string_list(plural, f"{name}.{plural_source_key}")
    raise ConfigError(f"{name}.{source_key} is required")


def _target_spreadsheet_id(
    item: Mapping[str, Any],
    name: str,
    default_spreadsheet_id: str | None,
) -> str:
    """Resolve the spreadsheet ID for a roster target."""
    value = item.get("spreadsheet_id", item.get("gsheet_id", default_spreadsheet_id))
    return _required_string(value, f"{name}.spreadsheet_id")


def _mapping(value: Any, name: str) -> Mapping[str, Any]:
    """Read a mapping config value."""
    if not isinstance(value, Mapping):
        raise ConfigError(f"{name} must be a mapping")
    return value


def _list(value: Any, name: str) -> list[Any]:
    """Read a list config value."""
    if not isinstance(value, list):
        raise ConfigError(f"{name} must be a list")
    return value


def _string_list(value: Any, name: str) -> list[str]:
    """Read a string list config value."""
    if not isinstance(value, list) or not all(isinstance(item, str) for item in value):
        raise ConfigError(f"{name} must be a list of strings")
    if not value:
        raise ConfigError(f"{name} must not be empty")
    return value


def _required_string(value: Any, name: str) -> str:
    """Read a required string config value."""
    if not isinstance(value, str) or not value:
        raise ConfigError(f"{name} must be a string")
    return value


def _optional_string(value: Any, name: str) -> str | None:
    """Read an optional string config value."""
    if value in (None, ""):
        return None
    return _required_string(value, name)


def _bool(value: Any, name: str) -> bool:
    """Read a boolean config value."""
    if not isinstance(value, bool):
        raise ConfigError(f"{name} must be a boolean")
    return value
