from __future__ import annotations

import datetime as dt
from pathlib import Path
from types import SimpleNamespace
from zoneinfo import ZoneInfo

import pytest
from openpyxl import load_workbook

from parishkit.config import ConfigError
from parishkit.google.auth import GoogleAPIError
from parishkit.google.drive import GOOGLE_SHEET_MIME_TYPE
from parishkit.parishsoft import ParishSoftData
from parishkit.pk_create_ps_ministry_rosters import (
    DRIVE_SCOPE,
    HEADER_BACKGROUND_COLOR,
    HEADER_TEXT_COLOR,
    ROSTER_COLUMN_WIDTHS,
    ROSTER_FROZEN_ROWS,
    ROSTER_TITLE_MERGE_COLUMNS,
    ROSTER_WORKSHEET_TITLE,
    RosterMember,
    load_drive_credentials,
    ministry_roster_members,
    roster_config_from_yaml,
    roster_drive_name,
    roster_role_matches,
    roster_values,
    roster_workbook,
    safe_roster_filename,
    workgroup_roster_members,
)
from parishkit.pk_create_ps_ministry_rosters import (
    main as create_ministry_rosters_main,
)


class Request:
    """Fake Google API request whose execute() returns a canned response."""

    def __init__(self, response=None, exc: Exception | None = None):
        """Store either the canned response or exception to raise."""
        self.response = {} if response is None else response
        self.exc = exc

    def execute(self):
        """Return the canned response unless configured to raise an error."""
        if self.exc is not None:
            raise self.exc
        return self.response


class DriveService:
    """Fake Drive service used as an identity token in roster tests."""


def test_drive_credentials_resolve_relative_paths(tmp_path, monkeypatch):
    """Relative Google Drive credential paths resolve against the config directory."""
    calls = []

    def fake_load(path, *, scopes, subject):
        """Capture the resolved service account path for assertion."""
        calls.append((path, scopes, subject))
        return object()

    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.load_service_account_credentials",
        fake_load,
    )

    load_drive_credentials(
        {
            "google": {
                "service_account_file": "credentials/google-service-account.json",
                "delegated_subject": "itadmin@example.org",
            }
        },
        base_dir=tmp_path,
    )

    assert calls[0][0] == tmp_path / "credentials" / "google-service-account.json"
    assert calls[0][1] == [DRIVE_SCOPE]


def write_config(tmp_path: Path, *, dry_run: bool = False) -> Path:
    """Write a rosters config (plus API key file) and return it.

    The config covers a ministry with a role sheet and a separate workgroup so
    one run exercises every roster-writing path. The setup stays local to this
    test so fixtures remain easy to understand and change.
    """
    api_key = tmp_path / "parishsoft-api-key.txt"
    api_key.write_text("key", encoding="utf-8")
    config = tmp_path / "config.yaml"
    config.write_text(
        f"""
common:
  dry_run: {str(dry_run).lower()}
parishsoft:
  api_key_file: {api_key}
  cache_dir: {tmp_path / "cache"}
  cache_limit: 1d
google:
  user_token_file: {tmp_path / "google-user-token.json"}
rosters:
  spreadsheet_id: default-sheet
  ministries:
    - ministry: Readers
      include_birthday: true
      role_sheets:
        - name: Reader Leads
          roles:
            - Lead
          spreadsheet_id: lead-sheet
  workgroups:
    - workgroup: Movers
      spreadsheet_id: movers-sheet
""",
        encoding="utf-8",
    )
    return config


def parishsoft_data() -> ParishSoftData:
    """Build a small ParishSoftData fixture with several roster members.

    Member 1 is a Readers Lead and Movers member with published contact info;
    member 2 is a Readers Member and a Movers leader (via the " Ldr" workgroup
    suffix) with unpublished contact info; member 3 has a blank ministry role;
    and member 4 has email but no phone. The setup stays local to this test so
    fixtures remain easy to understand and change.
    """
    family = {
        "familyDUID": 10,
        "primaryAddress1": "1 Main St",
        "primaryCity": "Louisville",
        "primaryState": "KY",
        "primaryPostalCode": "40202",
    }
    members = {
        1: {
            "memberDUID": 1,
            "familyDUID": 10,
            "firstName": "Ann",
            "lastName": "Smith",
            "py friendly name LF": "Smith, Ann",
            "py family": family,
            "family_PublishPhone": True,
            "family_PublishEMail": True,
            "mobilePhone": "502-555-1000",
            "py emailAddresses": ["ann@example.org"],
            "birthdate": dt.date(1980, 5, 4),
            "py ministries": {"Readers": {"role": "Lead"}},
            "py workgroups": {"Movers": {"name": "Movers"}},
        },
        2: {
            "memberDUID": 2,
            "familyDUID": 10,
            "firstName": "Bob",
            "lastName": "Adams",
            "py friendly name LF": "Adams, Bob",
            "py family": family,
            "family_PublishPhone": False,
            "family_PublishEMail": False,
            "py ministries": {"Readers": {"role": "Member"}},
            "py workgroups": {"Movers Ldr": {"name": "Movers Ldr"}},
        },
        3: {
            "memberDUID": 3,
            "familyDUID": 10,
            "firstName": "Chris",
            "lastName": "Role",
            "py friendly name LF": "Role, Chris",
            "py family": family,
            "family_PublishPhone": False,
            "family_PublishEMail": False,
            "py ministries": {"Readers": {"role": ""}},
            "py workgroups": {},
        },
        4: {
            "memberDUID": 4,
            "familyDUID": 10,
            "firstName": "Erin",
            "lastName": "Email",
            "py friendly name LF": "Email, Erin",
            "py family": family,
            "family_PublishPhone": True,
            "family_PublishEMail": True,
            "py emailAddresses": ["erin@example.org"],
            "py ministries": {"Readers": {"role": "Member"}},
            "py workgroups": {},
        },
    }
    return ParishSoftData(
        organization_id=7,
        families={10: family},
        members=members,
        family_groups={},
        family_workgroups={},
        family_workgroup_memberships={},
        member_contactinfos={},
        member_workgroups={},
        member_workgroup_memberships={},
        ministry_types={},
        ministry_type_memberships={},
        funds={},
        pledges={},
        contributions={},
    )


def fixed_update_time() -> dt.datetime:
    """Return a deterministic roster update timestamp for tests."""
    return dt.datetime(
        2026,
        1,
        2,
        3,
        4,
        tzinfo=ZoneInfo("America/Kentucky/Louisville"),
    )


def install_upload_recorder(monkeypatch):
    """Patch Drive uploads and capture workbook content before temp cleanup."""
    uploads = []

    def fake_upload(_service, file_id, xlsx_path, *, name):
        """Record the Drive file ID, upload name, and generated workbook."""
        workbook = load_workbook(xlsx_path)
        worksheet = workbook.active
        rows = [
            [cell.value for cell in row]
            for row in worksheet.iter_rows(max_row=worksheet.max_row)
        ]
        uploads.append(
            {
                "file_id": file_id,
                "name": name,
                "worksheet_title": worksheet.title,
                "rows": rows,
                "freeze_panes": worksheet.freeze_panes,
                "merged_ranges": {
                    str(range_) for range_ in worksheet.merged_cells.ranges
                },
                "widths": [
                    worksheet.column_dimensions[chr(ord("A") + index)].width
                    for index in range(len(ROSTER_COLUMN_WIDTHS))
                ],
            }
        )
        workbook.close()
        return {"id": file_id, "name": name}

    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.update_file_with_xlsx",
        fake_upload,
    )
    return uploads


def install_drive_preflight(monkeypatch, *, metadata_by_id=None, errors_by_id=None):
    """Patch Drive metadata lookups and capture preflighted file IDs."""
    calls = []
    metadata_by_id = {} if metadata_by_id is None else metadata_by_id
    errors_by_id = {} if errors_by_id is None else errors_by_id

    def fake_get_file_metadata(_service, file_id, *, fields):
        """Return canned Drive metadata or raise the configured API error."""
        calls.append((file_id, fields))
        if file_id in errors_by_id:
            raise errors_by_id[file_id]
        return metadata_by_id.get(
            file_id,
            {
                "id": file_id,
                "name": f"Sheet {file_id}",
                "mimeType": GOOGLE_SHEET_MIME_TYPE,
                "capabilities": {
                    "canEdit": True,
                    "canModifyContent": True,
                    "canRename": True,
                },
            },
        )

    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.get_file_metadata",
        fake_get_file_metadata,
    )
    return calls


def test_roster_config_validation_and_role_sheets():
    """Config parsing keeps the ministry name, multiple source ministries, and
    nested role-sheet targets (including the legacy "role sheets" key)."""
    config = roster_config_from_yaml(
        {
            "rosters": {
                "spreadsheet_id": "default-sheet",
                "ministries": [
                    {
                        "ministries": ["Readers", "Greeters"],
                        "name": "Welcome Ministers",
                        "include_birthday": True,
                        "role sheets": [
                            {
                                "name": "Leads",
                                "roles": ["Lead"],
                                "spreadsheet_id": "lead-sheet",
                            }
                        ],
                    }
                ],
            }
        }
    )

    assert config.ministries[0].name == "Welcome Ministers"
    assert config.ministries[0].source_names == ("Readers", "Greeters")
    assert config.ministries[0].role_sheets[0].spreadsheet_id == "lead-sheet"


def test_roster_config_rejects_missing_targets():
    """Config with neither ministries nor workgroups raises ConfigError."""
    with pytest.raises(ConfigError, match="ministries or workgroups"):
        roster_config_from_yaml({"rosters": {"ministries": []}})


def test_roster_config_rejects_unknown_target_key():
    """Misspelled roster target keys fail instead of being ignored."""
    with pytest.raises(ConfigError, match="unsupported key"):
        roster_config_from_yaml(
            {
                "rosters": {
                    "spreadsheet_id": "sheet-id",
                    "ministries": [
                        {
                            "ministry": "Readers",
                            "clear_rng": "Readers!A:Z",
                        }
                    ],
                }
            }
        )


def test_roster_config_rejects_removed_range_keys():
    """The XLSX upload writer rejects obsolete range settings."""
    with pytest.raises(ConfigError, match="unsupported key.*range"):
        roster_config_from_yaml(
            {
                "rosters": {
                    "spreadsheet_id": "default-sheet",
                    "range": "Roster!A1",
                    "ministries": [{"ministry": "Readers"}],
                }
            }
        )

    with pytest.raises(ConfigError, match="unsupported key.*clear_range"):
        roster_config_from_yaml(
            {
                "rosters": {
                    "spreadsheet_id": "default-sheet",
                    "ministries": [
                        {
                            "ministry": "Readers",
                            "clear_range": "Roster!A:Z",
                        }
                    ],
                }
            }
        )


def test_roster_config_requires_role_sheet_drive_file():
    """Role sheets must name their own Drive file because uploads replace files."""
    with pytest.raises(ConfigError, match=r"role_sheets\[0\]\.spreadsheet_id"):
        roster_config_from_yaml(
            {
                "rosters": {
                    "spreadsheet_id": "default-sheet",
                    "ministries": [
                        {
                            "ministry": "Readers",
                            "role_sheets": [
                                {
                                    "name": "Reader Leads",
                                    "roles": ["Lead"],
                                }
                            ],
                        }
                    ],
                }
            }
        )


def test_roster_config_rejects_workgroup_role_sheets():
    """Workgroup targets reject role-sheet config instead of ignoring it."""
    for key in ("role_sheets", "role sheets"):
        with pytest.raises(ConfigError, match="unsupported key"):
            roster_config_from_yaml(
                {
                    "rosters": {
                        "spreadsheet_id": "default-sheet",
                        "workgroups": [
                            {
                                "workgroup": "Movers",
                                key: [
                                    {
                                        "name": "Mover Leaders",
                                        "roles": ["Leader"],
                                        "spreadsheet_id": "leader-sheet",
                                    }
                                ],
                            }
                        ],
                    }
                }
            )


def test_roster_config_rejects_duplicate_drive_files():
    """Roster outputs must not share Drive files that would overwrite each other."""
    with pytest.raises(ConfigError, match="must not share"):
        roster_config_from_yaml(
            {
                "rosters": {
                    "spreadsheet_id": "default-sheet",
                    "ministries": [
                        {
                            "ministry": "Readers",
                            "role_sheets": [
                                {
                                    "name": "Reader Leads",
                                    "roles": ["Lead"],
                                    "spreadsheet_id": "default-sheet",
                                }
                            ],
                        }
                    ],
                }
            }
        )


def test_roster_generation_for_ministries_and_workgroups():
    """Roster members are sorted by name, roles resolve from ministry data and
    the workgroup leader suffix, and role matching is case/list aware."""
    data = parishsoft_data()

    ministry_members = ministry_roster_members(data, ["Readers"])
    workgroup_members = workgroup_roster_members(
        data,
        "Movers",
        leader_suffix=" Ldr",
    )

    assert [(item.member["memberDUID"], item.role) for item in ministry_members] == [
        (2, "Member"),
        (4, "Member"),
        (3, ""),
        (1, "Lead"),
    ]
    assert [(item.member["memberDUID"], item.role) for item in workgroup_members] == [
        (2, "Leader"),
        (1, "Member"),
    ]
    assert roster_role_matches("Lead, Member", {"Lead"})


def test_roster_values_separate_phone_and_email_rows():
    """roster_values separates email from phone when both contact types exist."""
    data = parishsoft_data()
    update_time = dt.datetime(
        2026,
        1,
        2,
        3,
        4,
        tzinfo=ZoneInfo("America/Kentucky/Louisville"),
    )

    values = roster_values(
        "Readers",
        [
            RosterMember(member=data.members[1], role="Lead"),
            RosterMember(member=data.members[4], role="Member"),
        ],
        include_birthday=True,
        now=update_time,
    )

    assert values[:4] == [
        ["Ministry: Readers"],
        ["Last updated: 2026-01-02 03:04:00 EST"],
        [],
        ["Member name", "Address", "Phone / email", "Birthday", "Role"],
    ]
    assert values[4] == [
        "Email, Erin",
        "1 Main St\nLouisville, KY 40202",
        "erin@example.org",
        "",
        "Member",
    ]
    assert values[5] == [
        "Smith, Ann",
        "1 Main St\nLouisville, KY 40202",
        "502-555-1000 cell",
        "May 4",
        "Lead",
    ]
    assert values[6] == ["", "", "ann@example.org", "", ""]


def test_roster_workbook_freezes_headers_and_sizes_columns():
    """Generated XLSX files freeze rows, style headers, and size columns."""
    workbook = roster_workbook(
        [
            ["Ministry: Readers"],
            ["Last updated: 2026-01-02 03:04:00 EST"],
            [],
            ["Member name", "Address", "Phone / email", "Birthday", "Role"],
            ["Smith, Ann", "1 Main St", "502-555-1000 cell", "May 4", "Lead"],
        ]
    )
    worksheet = workbook.active

    assert worksheet.title == ROSTER_WORKSHEET_TITLE
    assert worksheet.freeze_panes == f"A{ROSTER_FROZEN_ROWS + 1}"
    assert {str(range_) for range_ in worksheet.merged_cells.ranges} == {
        f"A1:{chr(ord('A') + ROSTER_TITLE_MERGE_COLUMNS - 1)}1",
        f"A2:{chr(ord('A') + ROSTER_TITLE_MERGE_COLUMNS - 1)}2",
    }
    assert worksheet["A1"].fill.fgColor.rgb.endswith(HEADER_BACKGROUND_COLOR)
    assert worksheet["A1"].font.color.rgb.endswith(HEADER_TEXT_COLOR)
    assert worksheet["A1"].alignment.horizontal == "left"
    assert worksheet["A4"].fill.fgColor.rgb.endswith(HEADER_BACKGROUND_COLOR)
    assert worksheet["A4"].font.color.rgb.endswith(HEADER_TEXT_COLOR)
    assert worksheet["A4"].alignment.horizontal == "center"
    assert worksheet["A5"].alignment.vertical == "top"
    assert worksheet["A5"].alignment.wrap_text is True
    assert [
        worksheet.column_dimensions[chr(ord("A") + index)].width
        for index in range(len(ROSTER_COLUMN_WIDTHS))
    ] == list(ROSTER_COLUMN_WIDTHS)
    workbook.close()


def test_roster_workbook_preserves_formula_like_text(tmp_path):
    """Generated XLSX files keep ParishSoft text from becoming formulas."""
    workbook_path = tmp_path / "roster.xlsx"
    workbook = roster_workbook([["=1+1"]])
    worksheet = workbook.active

    assert worksheet["A1"].value == "=1+1"
    assert worksheet["A1"].data_type == "s"

    workbook.save(workbook_path)
    workbook.close()
    reloaded = load_workbook(workbook_path, data_only=False)

    assert reloaded.active["A1"].value == "=1+1"
    assert reloaded.active["A1"].data_type == "s"
    reloaded.close()


def test_roster_drive_name_sanitizes_slashes_and_uses_timestamp():
    """Drive names are readable and avoid slash-separated path fragments."""
    update_time = dt.datetime(
        2026,
        1,
        2,
        3,
        4,
        tzinfo=ZoneInfo("America/Kentucky/Louisville"),
    )

    assert safe_roster_filename("Care / Prayer") == "Care - Prayer"
    assert roster_drive_name("Care / Prayer", update_time) == (
        "Care - Prayer as of 2026-01-02 03:04:00 EST"
    )


def test_create_ministry_rosters_main_uploads_workbooks(
    tmp_path,
    monkeypatch,
    capsys,
):
    """main uploads one XLSX workbook per configured roster target.

    The ParishSoft client is stubbed out so no real data is loaded, and the
    loader is replaced with a recorder to confirm the expected load options.
    The setup stays local to this test so fixtures remain easy to understand
    and change.
    """
    service = DriveService()
    uploads = install_upload_recorder(monkeypatch)
    preflight_calls = install_drive_preflight(monkeypatch)
    loader_calls = []
    # Avoid building a real ParishSoft client; the loader returns canned data.
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.parishsoft_client_from_config",
        lambda _common, _config: SimpleNamespace(),
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.current_roster_time",
        lambda _timezone_name: fixed_update_time(),
    )

    def loader(_client, **kwargs):
        """Record loader kwargs and return the canned ParishSoft fixture."""
        loader_calls.append(kwargs)
        return parishsoft_data()

    assert (
        create_ministry_rosters_main(
            ["--config", str(write_config(tmp_path)), "--debug"],
            loader=loader,
            drive_factory=lambda _config: service,
        )
        == 0
    )

    error = capsys.readouterr().err
    assert "Ministry roster operation completed successfully" in error
    assert "Ministry roster targets: Readers" in error
    assert "Workgroup roster targets: Movers" in error
    assert "Uploaded roster workbook for Readers to Drive file default-sheet" in error
    assert "Uploaded roster workbook Readers as of" not in error
    assert "RosterTarget(" not in error
    assert loader_calls == [{"active_only": True, "parishioners_only": False}]
    assert [upload["file_id"] for upload in uploads] == [
        "default-sheet",
        "lead-sheet",
        "movers-sheet",
    ]
    assert [call[0] for call in preflight_calls] == [
        "default-sheet",
        "lead-sheet",
        "movers-sheet",
    ]
    assert [upload["name"] for upload in uploads] == [
        "Readers as of 2026-01-02 03:04:00 EST",
        "Reader Leads as of 2026-01-02 03:04:00 EST",
        "Movers as of 2026-01-02 03:04:00 EST",
    ]
    assert uploads[0]["worksheet_title"] == ROSTER_WORKSHEET_TITLE
    assert uploads[0]["freeze_panes"] == "A5"
    assert uploads[0]["merged_ranges"] == {"A1:D1", "A2:D2"}
    assert uploads[0]["widths"] == list(ROSTER_COLUMN_WIDTHS)
    assert uploads[0]["rows"][0][0] == "Ministry: Readers"
    assert uploads[0]["rows"][1][0] == "Last updated: 2026-01-02 03:04:00 EST"
    assert uploads[0]["rows"][4][4] == "Member"
    assert uploads[0]["rows"][6][4] is None
    assert uploads[1]["rows"][4][0] == "Smith, Ann"
    assert uploads[2]["rows"][4][0] == "Adams, Bob"


def test_create_ministry_rosters_dry_run_skips_drive_uploads(
    tmp_path,
    monkeypatch,
    capsys,
):
    """In dry-run mode main loads data but performs no Drive uploads."""
    service = DriveService()
    uploads = install_upload_recorder(monkeypatch)
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.parishsoft_client_from_config",
        lambda _common, _config: SimpleNamespace(),
    )

    assert (
        create_ministry_rosters_main(
            ["--config", str(write_config(tmp_path, dry_run=True))],
            loader=lambda _client, **_kwargs: parishsoft_data(),
            drive_factory=lambda _config: service,
        )
        == 0
    )

    assert uploads == []
    error = capsys.readouterr().err
    assert "dry-run: would upload" in error
    assert "for Readers to Drive file default-sheet" in error
    assert "would upload 8 row(s) as Readers as of" not in error


def test_create_ministry_rosters_upload_failure_stops_later_uploads(
    tmp_path,
    monkeypatch,
):
    """A failed Drive upload aborts before later targets are uploaded."""
    service = DriveService()
    uploads = []
    install_drive_preflight(monkeypatch)

    def fake_upload(_service, file_id, _xlsx_path, *, name):
        """Record the attempted upload, then fail like Drive did."""
        uploads.append((file_id, name))
        raise GoogleAPIError(500, "temporary Drive failure")

    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.update_file_with_xlsx",
        fake_upload,
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.current_roster_time",
        lambda _timezone_name: fixed_update_time(),
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.parishsoft_client_from_config",
        lambda _common, _config: SimpleNamespace(),
    )

    assert (
        create_ministry_rosters_main(
            ["--config", str(write_config(tmp_path))],
            loader=lambda _client, **_kwargs: parishsoft_data(),
            drive_factory=lambda _config: service,
        )
        == 2
    )

    assert uploads == [("default-sheet", "Readers as of 2026-01-02 03:04:00 EST")]


def test_create_ministry_rosters_prepares_all_workbooks_before_uploads(
    tmp_path,
    monkeypatch,
    capsys,
):
    """Local XLSX build failures abort before any Drive file is replaced."""
    service = DriveService()
    uploads = install_upload_recorder(monkeypatch)
    install_drive_preflight(monkeypatch)
    workbook_calls = {"count": 0}

    def fake_roster_workbook(values):
        """Fail on the second workbook to simulate bad local XLSX data."""
        workbook_calls["count"] += 1
        if workbook_calls["count"] == 2:
            raise ValueError("bad local workbook data")
        return roster_workbook(values)

    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.roster_workbook",
        fake_roster_workbook,
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.current_roster_time",
        lambda _timezone_name: fixed_update_time(),
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.parishsoft_client_from_config",
        lambda _common, _config: SimpleNamespace(),
    )

    assert (
        create_ministry_rosters_main(
            ["--config", str(write_config(tmp_path))],
            loader=lambda _client, **_kwargs: parishsoft_data(),
            drive_factory=lambda _config: service,
        )
        == 2
    )

    error = capsys.readouterr().err
    assert "Could not build XLSX roster workbook for 'Reader Leads'" in error
    assert "bad local workbook data" in error
    assert uploads == []


def test_create_ministry_rosters_preflight_failure_skips_uploads(
    tmp_path,
    monkeypatch,
    capsys,
):
    """A bad Drive target fails before any configured file is replaced."""
    service = DriveService()
    uploads = install_upload_recorder(monkeypatch)
    install_drive_preflight(
        monkeypatch,
        metadata_by_id={
            "movers-sheet": {
                "id": "movers-sheet",
                "name": "Movers",
                "mimeType": GOOGLE_SHEET_MIME_TYPE,
                "capabilities": {"canEdit": False},
            },
        },
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.current_roster_time",
        lambda _timezone_name: fixed_update_time(),
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.parishsoft_client_from_config",
        lambda _common, _config: SimpleNamespace(),
    )

    assert (
        create_ministry_rosters_main(
            ["--config", str(write_config(tmp_path))],
            loader=lambda _client, **_kwargs: parishsoft_data(),
            drive_factory=lambda _config: service,
        )
        == 2
    )

    error = capsys.readouterr().err
    assert "not editable by the delegated Google user" in error
    assert "movers-sheet" in error
    assert uploads == []


def test_create_ministry_rosters_preflight_requires_content_changes(
    tmp_path,
    monkeypatch,
    capsys,
):
    """A Drive target must allow replacing file content before uploads begin."""
    service = DriveService()
    uploads = install_upload_recorder(monkeypatch)
    install_drive_preflight(
        monkeypatch,
        metadata_by_id={
            "lead-sheet": {
                "id": "lead-sheet",
                "name": "Lead Sheet",
                "mimeType": GOOGLE_SHEET_MIME_TYPE,
                "capabilities": {"canEdit": True, "canModifyContent": False},
            },
        },
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.current_roster_time",
        lambda _timezone_name: fixed_update_time(),
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.parishsoft_client_from_config",
        lambda _common, _config: SimpleNamespace(),
    )

    assert (
        create_ministry_rosters_main(
            ["--config", str(write_config(tmp_path))],
            loader=lambda _client, **_kwargs: parishsoft_data(),
            drive_factory=lambda _config: service,
        )
        == 2
    )

    error = capsys.readouterr().err
    assert "cannot have its content replaced" in error
    assert "lead-sheet" in error
    assert uploads == []


def test_create_ministry_rosters_preflight_requires_rename_permission(
    tmp_path,
    monkeypatch,
    capsys,
):
    """A Drive target must allow renaming before uploads begin."""
    service = DriveService()
    uploads = install_upload_recorder(monkeypatch)
    install_drive_preflight(
        monkeypatch,
        metadata_by_id={
            "movers-sheet": {
                "id": "movers-sheet",
                "name": "Movers",
                "mimeType": GOOGLE_SHEET_MIME_TYPE,
                "capabilities": {
                    "canEdit": True,
                    "canModifyContent": True,
                    "canRename": False,
                },
            },
        },
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.current_roster_time",
        lambda _timezone_name: fixed_update_time(),
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.parishsoft_client_from_config",
        lambda _common, _config: SimpleNamespace(),
    )

    assert (
        create_ministry_rosters_main(
            ["--config", str(write_config(tmp_path))],
            loader=lambda _client, **_kwargs: parishsoft_data(),
            drive_factory=lambda _config: service,
        )
        == 2
    )

    error = capsys.readouterr().err
    assert "cannot be renamed by the delegated Google user" in error
    assert "movers-sheet" in error
    assert uploads == []


def test_create_ministry_rosters_dry_run_does_not_require_google_config(
    tmp_path,
    monkeypatch,
):
    """Dry-run roster generation does not build a Google Drive service."""
    config = write_config(tmp_path, dry_run=True)
    text = config.read_text(encoding="utf-8")
    config.write_text(
        text.replace(
            f"google:\n  user_token_file: {tmp_path / 'google-user-token.json'}\n",
            "",
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.parishsoft_client_from_config",
        lambda _common, _config: SimpleNamespace(),
    )

    assert (
        create_ministry_rosters_main(
            ["--config", str(config)],
            loader=lambda _client, **_kwargs: parishsoft_data(),
        )
        == 0
    )


def test_create_ministry_rosters_reports_missing_parishsoft_source(
    tmp_path,
    monkeypatch,
    capsys,
):
    """Unknown roster ministry/workgroup source names abort before Drive uploads."""
    service = DriveService()
    uploads = install_upload_recorder(monkeypatch)
    config = write_config(tmp_path)
    text = config.read_text(encoding="utf-8")
    config.write_text(
        text.replace("Readers", "Missing Readers").replace(
            "Movers",
            "Missing Movers",
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.parishsoft_client_from_config",
        lambda _common, _config: SimpleNamespace(),
    )

    assert (
        create_ministry_rosters_main(
            ["--config", str(config)],
            loader=lambda _client, **_kwargs: parishsoft_data(),
            drive_factory=lambda _config: service,
        )
        == 2
    )

    error = capsys.readouterr().err
    assert "ERROR parishkit.pk_create_ps_ministry_rosters" in error
    assert "Configured ParishSoft ministry was not found" in error
    assert "rosters.ministries[].ministry" in error
    assert uploads == []


def test_create_ministry_rosters_validation_uses_custom_leader_suffix(
    tmp_path,
    monkeypatch,
):
    """A leader-only workgroup source can use the configured suffix."""
    service = DriveService()
    uploads = install_upload_recorder(monkeypatch)
    install_drive_preflight(monkeypatch)
    config = write_config(tmp_path)
    config.write_text(
        config.read_text(encoding="utf-8").replace(
            "rosters:\n",
            'rosters:\n  workgroup_leader_suffix: " Captain"\n',
        ),
        encoding="utf-8",
    )
    data = parishsoft_data()
    data.members[1]["py workgroups"] = {}
    data.members[2]["py workgroups"] = {"Movers Captain": {"name": "Movers Captain"}}
    monkeypatch.setattr(
        "parishkit.pk_create_ps_ministry_rosters.parishsoft_client_from_config",
        lambda _common, _config: SimpleNamespace(),
    )

    assert (
        create_ministry_rosters_main(
            ["--config", str(config)],
            loader=lambda _client, **_kwargs: data,
            drive_factory=lambda _config: service,
        )
        == 0
    )

    assert uploads[2]["rows"][4][0] == "Adams, Bob"
    assert uploads[2]["rows"][4][3] == "Leader"


def test_create_ministry_rosters_reports_invalid_yaml(tmp_path, capsys):
    """Invalid YAML exits cleanly with a repair-oriented error message."""
    config = tmp_path / "config.yaml"
    config.write_text("common:\n  dry_run: true\n  bad: [\n", encoding="utf-8")

    assert create_ministry_rosters_main(["--config", str(config)]) == 2

    error = capsys.readouterr().err
    assert "ERROR: could not parse YAML config file" in error
    assert "Check indentation" in error


def test_create_ministry_rosters_logs_config_validation_error(tmp_path, capsys):
    """Roster config validation failures are logged as ERROR before exit."""
    config = tmp_path / "config.yaml"
    config.write_text(
        "common:\n  dry_run: true\nrosters:\n  ministries: []\n",
        encoding="utf-8",
    )

    assert create_ministry_rosters_main(["--config", str(config)]) == 2

    error = capsys.readouterr().err
    assert "ERROR parishkit.pk_create_ps_ministry_rosters" in error
    assert "Configuration validation failed" in error
    assert "rosters must configure ministries or workgroups" in error
